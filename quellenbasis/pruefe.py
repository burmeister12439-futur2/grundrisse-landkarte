#!/usr/bin/env python3
"""Prueft die Quellenbasis der GRUNDRISSE und meldet ihren Stand.

Zwei harte Regeln, sonst nur Bericht. Aufruf aus dem Wurzelverzeichnis:
    python3 quellenbasis/pruefe.py
"""
import os, re, sys, glob, datetime
from openpyxl import load_workbook

WURZEL = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
fehler = []
hinweise = []

def neuestes_register():
    d = sorted(glob.glob(os.path.join(WURZEL, "quellenbasis", "register", "*_Grundrisse_Foresight-Register_2045_v*.xlsx")))
    if not d:
        print("FEHLER: kein Register gefunden"); sys.exit(1)
    def num(p):
        m = re.search(r"_v(\d+)\.xlsx$", p); return int(m.group(1)) if m else 0
    return max(d, key=num)

pfad = neuestes_register()
wb = load_workbook(pfad)
ws = wb["Schlüsselquellen (Grundrisse)"]
kopf = [c.value for c in ws[1]]
def sp(name):
    return kopf.index(name) + 1 if name in kopf else None
iNr, iTitel, iVer = sp("Nr"), sp("Titel"), sp("Verifikationsstand")
iDatei, iFund = sp("Datei (Material/Themen)"), sp("Fundort / Quelle")

nummern, zeilen, mit_datei = set(), 0, 0
for r in range(2, ws.max_row + 1):
    nr = ws.cell(row=r, column=iNr).value
    if nr in (None, ""): continue
    zeilen += 1
    nummern.add(str(nr).strip())
    datei = ws.cell(row=r, column=iDatei).value if iDatei else None
    fund = str(ws.cell(row=r, column=iFund).value or "")
    beleg = bool(datei) or "http" in fund.lower()
    if beleg: mit_datei += 1
    ver = str(ws.cell(row=r, column=iVer).value or "")
    # Regel 1
    if ver.startswith("verifiziert") and not beleg:
        fehler.append("Zeile %s '%s' gilt als verifiziert, hat aber weder Datei noch Link"
                      % (nr, str(ws.cell(row=r, column=iTitel).value)[:60]))

# Regel 2 und 3: zitierte Registernummern
# Nur Herkunftszeilen auswerten, nicht den eingebetteten Registerdatensatz.
iGuete = sp("Güte")
guete = {}
for r in range(2, ws.max_row + 1):
    nr = ws.cell(row=r, column=iNr).value
    if nr in (None, ""): continue
    guete[str(nr).strip()] = str(ws.cell(row=r, column=iGuete).value or "?") if iGuete else "?"

BELEGFAEHIG = ("A", "B", "C", "F")
zitiert = set()
for html in glob.glob(os.path.join(WURZEL, "*.html")):
    with open(html, encoding="utf-8", errors="replace") as fh:
        text = fh.read()
    for zeile in re.findall(r'class="herkrow">(.*?)</div>', text, re.S):
        klar = re.sub(r"<[^>]+>", " ", zeile)
        for m in re.finditer(r"Nr\.?\s*(\d{1,3})(?:\s+und\s+(\d{1,3}))?", klar):
            zitiert.add(m.group(1))
            if m.group(2): zitiert.add(m.group(2))
for z in sorted(zitiert, key=int):
    if z not in nummern:
        fehler.append("Registernummer %s wird auf der Seite zitiert, steht aber nicht im Register" % z)
        continue
    g = guete.get(z, "?")
    if g not in BELEGFAEHIG:
        fehler.append("Nr %s hat Guete %s und darf nichts belegen. Siehe QUELLENSTANDARD.md" % (z, g))
    elif g == "F":
        hinweise.append("Nr %s ist eigenes Werk (F). Der Pruefer kann nicht sehen, ob die Aussage "
                        "von der eigenen Position handelt. Das bleibt redaktionelle Pruefung." % z)

ausz = len([f for f in glob.glob(os.path.join(WURZEL, "quellenbasis", "auszuege", "*")) if os.path.isfile(f)])
m = re.search(r"(\d{4})-(\d{2})-(\d{2})", os.path.basename(pfad))
alter = None
if m:
    alter = (datetime.date.today() - datetime.date(*map(int, m.groups()))).days

print("Quellenbasis GRUNDRISSE")
print("  Register:            %s" % os.path.basename(pfad))
print("  Zeilen:              %d" % zeilen)
print("  mit Datei oder Link: %d" % mit_datei)
print("  zitiert auf Seite:   %d" % len(zitiert))
print("  Auszuege:             %d" % ausz)
if alter is not None:
    print("  Registerfassung:     %d Tage alt" % alter)
    if alter > 120:
        print("  HINWEIS: seit mehr als einem Quartal kein Zuwachs. Quartalslauf pruefen.")

if hinweise:
    print("\nHINWEISE (%d), keine Fehler:" % len(hinweise))
    for h in hinweise: print("  -", h)

if fehler:
    print("\nFEHLER (%d):" % len(fehler))
    for f in fehler: print("  -", f)
    sys.exit(1)
print("\nAlles in Ordnung.")
